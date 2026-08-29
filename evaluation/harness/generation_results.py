"""Result writer for generation evaluation outputs."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from evaluation.harness.generation_schema import GenerationRunMeta


class GenerationRunWriter:
    """Persists one generation evaluation run."""

    def __init__(self, results_dir: Path, run_meta: GenerationRunMeta) -> None:
        self.results_dir = results_dir
        self.run_meta = run_meta
        self.run_dir = self.results_dir / self.run_meta.run_id
        self.run_dir.mkdir(parents=True, exist_ok=False)

    def write(self, results_rows: list[dict[str, Any]]) -> Path:
        self._write_run_meta()
        self._write_results_workbook(self.run_dir / "results.xlsx", results_rows)
        return self.run_dir

    def _write_run_meta(self) -> None:
        (self.run_dir / "run_meta.json").write_text(
            json.dumps(self.run_meta.to_dict(), indent=2) + "\n",
            encoding="utf-8",
        )

    def _write_results_workbook(self, path: Path, rows: list[dict[str, Any]]) -> None:
        wb = Workbook()
        sheet = wb.active
        sheet.title = "results"
        self._write_sheet(sheet, rows)
        self._write_sheet(wb.create_sheet("category_results"), self._aggregate(rows, group_key="category"))
        self._write_sheet(wb.create_sheet("difficulty_results"), self._aggregate(rows, group_key="difficulty"))
        self._write_sheet(wb.create_sheet("overall_summary"), self._aggregate(rows, group_key=None))
        self._write_sheet(wb.create_sheet("failures"), self._failures(rows))
        wb.save(path)

    @staticmethod
    def _write_sheet(sheet: Any, rows: list[dict[str, Any]]) -> None:
        if not rows:
            return
        headers = list(rows[0].keys())
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header) for header in headers])

    @staticmethod
    def _aggregate(rows: list[dict[str, Any]], *, group_key: str | None) -> list[dict[str, Any]]:
        numeric_fields = (
            "correctness_score",
            "completeness_score",
            "abstention_quality_score",
            "faithfulness_score",
            "citation_precision",
            "citation_recall_proxy",
            "gold_service_overlap",
            "gold_file_overlap",
            "latency_ms",
            "retrieval_latency_ms",
            "generation_latency_ms",
        )
        buckets: dict[tuple[str, str, str], dict[str, Any]] = {}
        for row in rows:
            mode = str(row.get("mode", ""))
            policy = str(row.get("decomposition_policy", ""))
            group_value = str(row.get(group_key, "")) if group_key else "all"
            key = (mode, policy, group_value)
            if key not in buckets:
                buckets[key] = {
                    "mode": mode,
                    "decomposition_policy": policy,
                    "query_count": 0,
                    "hallucination_count": 0.0,
                    "error_count": 0.0,
                    **{field: 0.0 for field in numeric_fields},
                }
                if group_key:
                    buckets[key][group_key] = group_value
            bucket = buckets[key]
            bucket["query_count"] += 1
            bucket["hallucination_count"] += float(row.get("hallucination_flag", 0) or 0)
            bucket["error_count"] += 1.0 if str(row.get("error", "")).strip() else 0.0
            for field in numeric_fields:
                bucket[field] += float(row.get(field, 0.0) or 0.0)

        aggregated: list[dict[str, Any]] = []
        for key in sorted(buckets):
            bucket = buckets[key]
            count = float(bucket["query_count"] or 1)
            out: dict[str, Any] = {
                "mode": bucket["mode"],
                "decomposition_policy": bucket["decomposition_policy"],
                "query_count": int(bucket["query_count"]),
                "hallucination_rate": bucket["hallucination_count"] / count,
                "error_rate": bucket["error_count"] / count,
            }
            if group_key:
                out[group_key] = bucket[group_key]
            for field in numeric_fields:
                out[field] = bucket[field] / count
            aggregated.append(out)
        return aggregated

    @staticmethod
    def _failures(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        failures: list[dict[str, Any]] = []
        for row in rows:
            if (
                float(row.get("correctness_score", 0.0) or 0.0) < 0.5
                or int(row.get("hallucination_flag", 0) or 0) > 0
                or str(row.get("error", "")).strip()
            ):
                failures.append(
                    {
                        "run_id": row.get("run_id"),
                        "query_id": row.get("query_id"),
                        "mode": row.get("mode"),
                        "decomposition_policy": row.get("decomposition_policy"),
                        "correctness_score": row.get("correctness_score"),
                        "completeness_score": row.get("completeness_score"),
                        "faithfulness_score": row.get("faithfulness_score"),
                        "hallucination_flag": row.get("hallucination_flag"),
                        "error": row.get("error"),
                        "generated_answer": str(row.get("generated_answer", ""))[:600],
                        "gold_answer": str(row.get("gold_answer", ""))[:600],
                        "judge_rationale": row.get("judge_rationale"),
                        "citations_count": row.get("citations_count"),
                    }
                )
        return failures
