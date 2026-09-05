"""Result writer for retrieval evaluation outputs."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from evaluation.harness.schema import RunMeta


class RunWriter:
    """Persists one retrieval evaluation run."""

    def __init__(self, results_dir: Path, run_meta: RunMeta) -> None:
        self.results_dir = results_dir
        self.run_meta = run_meta
        self.run_dir = self.results_dir / self.run_meta.run_id
        self.run_dir.mkdir(parents=True, exist_ok=False)

    def write(self, results_rows: list[dict[str, Any]]) -> Path:
        self._write_run_meta()
        self._write_results_workbook(self.run_dir / "results.xlsx", results_rows)
        return self.run_dir

    def _write_run_meta(self) -> None:
        (self.run_dir / "run_meta.json").write_text(
            json.dumps(self.run_meta.to_dict(), indent=2) + "\n",
            encoding="utf-8",
        )

    def _write_results_workbook(self, path: Path, results_rows: list[dict[str, Any]]) -> None:
        normalized_rows = [self._normalize_result_row(row) for row in results_rows]

        workbook = Workbook()
        results_sheet = workbook.active
        results_sheet.title = "results"
        self._write_sheet(results_sheet, normalized_rows)

        category_rows = self._aggregate_rows(normalized_rows, group_key="category")
        category_sheet = workbook.create_sheet("category_results")
        self._write_sheet(category_sheet, category_rows)

        difficulty_rows = self._aggregate_rows(normalized_rows, group_key="difficulty")
        difficulty_sheet = workbook.create_sheet("difficulty_results")
        self._write_sheet(difficulty_sheet, difficulty_rows)

        workbook.save(path)

    @staticmethod
    def _write_sheet(sheet: Any, rows: list[dict[str, Any]]) -> None:
        if not rows:
            return
        fieldnames = list(rows[0].keys())
        sheet.append(fieldnames)
        for row in rows:
            sheet.append([row.get(field) for field in fieldnames])

    @staticmethod
    def _aggregate_rows(results_rows: list[dict[str, Any]], *, group_key: str) -> list[dict[str, Any]]:
        metrics = (
            "recall",
            "precision",
            "f1",
            "mrr",
            "ndcg",
            "recall_ceiling",
            "ceiling_adjusted_recall",
            "evidence_group_recall",
            "hit_count",
            "evidence_group_hit_count",
            "latency_ms",
        )
        groups: dict[tuple[str, int, str], dict[str, Any]] = {}

        for row in results_rows:
            strategy = str(row.get("strategy", ""))
            k = int(row.get("k", 0) or 0)
            group_value = str(row.get(group_key, ""))
            key = (strategy, k, group_value)
            if key not in groups:
                groups[key] = {
                    "strategy": strategy,
                    "k": k,
                    group_key: group_value,
                    "n": 0,
                    "recall": 0.0,
                    "precision": 0.0,
                    "f1": 0.0,
                    "mrr": 0.0,
                    "ndcg": 0.0,
                    "recall_ceiling": 0.0,
                    "ceiling_adjusted_recall": 0.0,
                    "evidence_group_recall": 0.0,
                    "hit_count": 0.0,
                    "evidence_group_hit_count": 0.0,
                    "latency_ms": 0.0,
                }
            bucket = groups[key]
            bucket["n"] += 1
            for metric in metrics:
                bucket[metric] += float(row.get(metric, 0.0))

        rows: list[dict[str, Any]] = []
        for _, aggregate in sorted(
            groups.items(),
            key=lambda item: (
                item[1]["strategy"],
                int(item[1]["k"]),
                int(item[1][group_key]) if group_key == "difficulty" and str(item[1][group_key]).isdigit() else str(item[1][group_key]),
            ),
        ):
            count = aggregate["n"]
            rows.append(
                {
                    "strategy": aggregate["strategy"],
                    "k": aggregate["k"],
                    group_key: aggregate[group_key],
                    "n": count,
                    "recall": aggregate["recall"] / count,
                    "precision": aggregate["precision"] / count,
                    "f1": aggregate["f1"] / count,
                    "mrr": aggregate["mrr"] / count,
                    "ndcg": aggregate["ndcg"] / count,
                    "recall_ceiling": aggregate["recall_ceiling"] / count,
                    "ceiling_adjusted_recall": aggregate["ceiling_adjusted_recall"] / count,
                    "evidence_group_recall": aggregate["evidence_group_recall"] / count,
                    "hit_count": aggregate["hit_count"] / count,
                    "evidence_group_hit_count": aggregate["evidence_group_hit_count"] / count,
                    "latency_ms": aggregate["latency_ms"] / count,
                }
            )
        return rows

    @staticmethod
    def _normalize_result_row(row: dict[str, Any]) -> dict[str, Any]:
        normalized = dict(row)
        if "latency_ms" not in normalized:
            normalized["latency_ms"] = normalized.get("graph_total_latency_ms", 0.0)
        return normalized
